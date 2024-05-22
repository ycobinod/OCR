import requests
import openpyxl

doc_url = 'https://<YOUR_ELASTIC_URL>/<YOUR_ELASTIC_INDEX>/_doc'

headers = { "Authorization": "ApiKey <YOUR_ELASTIC_API_KEY>"}

excel_file = openpyxl.load_workbook("static/files/books_list.xlsx")
sheet = excel_file.active

for i in range(2 , sheet.max_row+1):
    dic = {}
    book_name = sheet.cell(row=i , column=2).value
    if book_name:
        dic['book_name'] = sheet.cell(row=i , column=2).value
        column_3 = sheet.cell(row=i , column=3).value
        if column_3 is not None :
            try:
                split_string = column_3.split('|') 
                dic['author'] = split_string[0]
                dic['published_at'] = split_string[1]
            except:
                pass
        dic['rating'] = sheet.cell(row=i , column=4).value
        dic['price'] = sheet.cell(row=i , column=6).value
        dic['image_url'] = sheet.cell(row=i , column=9).value
        dic['amazon_id'] = sheet.cell(row=i , column=8).value
        dic['product_url'] = sheet.cell(row=i , column=1).value
        try:
            upload_resp = requests.post(doc_url, headers=headers, json=dic , verify=False)
            print('upload res----' , upload_resp)
        except Exception as err:
            print('er----' , err)